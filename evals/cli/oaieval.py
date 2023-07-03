"""
This file defines the `oaieval` CLI for running evals.
"""
import argparse
import logging
import shlex
import sys
from typing import Any, Mapping, Optional

import openai

import evals
import evals.api
import evals.base
import evals.record
from evals.registry import Registry

import openpyxl

from openpyxl import Workbook
from openpyxl import load_workbook


logger = logging.getLogger(__name__)


def _purple(str):
    return f"\033[1;35m{str}\033[0m"


def get_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run evals through the API")
    parser.add_argument(
        "completion_fn",
        type=str,
        help="One or more CompletionFn URLs, separated by commas (,). A CompletionFn can either be the name of a model available in the OpenAI API or a key in the registry (see evals/registry/completion_fns).",
    )
    parser.add_argument("eval", type=str, help="Name of an eval. See registry.")
    parser.add_argument("--extra_eval_params", type=str, default="")
    parser.add_argument("--max_samples", type=int, default=None)
    parser.add_argument("--cache", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--visible", action=argparse.BooleanOptionalAction, default=None)
    parser.add_argument("--seed", type=int, default=20220722)
    parser.add_argument("--user", type=str, default="")
    parser.add_argument("--record_path", type=str, default=None)
    parser.add_argument(
        "--log_to_file", type=str, default=None, help="Log to a file instead of stdout"
    )
    parser.add_argument(
        "--registry_path", type=str, default=None, action="append", help="Path to the registry"
    )
    parser.add_argument("--debug", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--local-run", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--dry-run", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--dry-run-logging", action=argparse.BooleanOptionalAction, default=True)
    return parser


def run(args, phrase, registry: Optional[Registry] = None):
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    visible = args.visible if args.visible is not None else (args.max_samples is None)

    if args.max_samples is not None:
        evals.eval.set_max_samples(args.max_samples)

    registry = registry or Registry()
    if args.registry_path:
        registry.add_registry_paths(args.registry_path)

    eval_spec = registry.get_eval(args.eval)
    assert (
        eval_spec is not None
    ), f"Eval {args.eval} not found. Available: {list(sorted(registry._evals.keys()))}"

    completion_fns = args.completion_fn.split(",")
    completion_fn_instances = [registry.make_completion_fn(url) for url in completion_fns]

    run_config = {
        "completion_fns": completion_fns,
        "eval_spec": eval_spec,
        "seed": args.seed,
        "max_samples": args.max_samples,
        "command": " ".join(map(shlex.quote, sys.argv)),
        "initial_settings": {
            "visible": visible,
        },
    }

    eval_name = eval_spec.key
    run_spec = evals.base.RunSpec(
        completion_fns=completion_fns,
        eval_name=eval_name,
        base_eval=eval_name.split(".")[0],
        split=eval_name.split(".")[1],
        run_config=run_config,
        created_by=args.user,
    )
    if args.record_path is None:
        record_path = f"/tmp/evallogs/{run_spec.run_id}_{args.completion_fn}_{args.eval}.jsonl"
    else:
        record_path = args.record_path
    if args.dry_run:
        recorder = evals.record.DummyRecorder(run_spec=run_spec, log=args.dry_run_logging)
    elif args.local_run:
        recorder = evals.record.LocalRecorder(record_path, run_spec=run_spec)
    else:
        recorder = evals.record.Recorder(record_path, run_spec=run_spec)

    api_extra_options = {}
    if not args.cache:
        api_extra_options["cache_level"] = 0

    run_url = f"{run_spec.run_id}"
    logger.info(_purple(f"Run started: {run_url}"))

    def parse_extra_eval_params(param_str: Optional[str]) -> Mapping[str, Any]:
        """Parse a string of the form "key1=value1,key2=value2" into a dict."""
        if not param_str:
            return {}

        def to_number(x):
            try:
                return int(x)
            except:
                pass
            try:
                return float(x)
            except:
                pass
            return x

        str_dict = dict(kv.split("=") for kv in param_str.split(","))
        return {k: to_number(v) for k, v in str_dict.items()}

    extra_eval_params = parse_extra_eval_params(args.extra_eval_params)

    eval_class = registry.get_class(eval_spec)
    eval = eval_class(
        completion_fns=completion_fn_instances,
        seed=args.seed,
        name=eval_name,
        registry=registry,
        **extra_eval_params,
    )
    result = eval.run(recorder, phrase)
    #print(result)
    return result

def testing(model, test, phrase):
    parser = get_parser()
    a =  [model, test]
    args = parser.parse_args(a)
    def ordenar_lista_tuplas(lista):
        lista_ordenada = sorted(lista, key=lambda tupla: tupla[1])
        return lista_ordenada
    result = ordenar_lista_tuplas(run(args, phrase))
    return result
#print(testing('gpt-3.5-turbo', 'tests.dev.v0', "Let's think step by step"))


def crear_tabla_excel(string, array):
    # Crea un nuevo libro de Excel y selecciona la hoja activa
    libro = Workbook()
    hoja = libro.active

    # Escribe el título en la primera columna de la primera fila
    hoja.cell(row=1, column=1).value = "Prompt"

    # Escribe los primeros elementos de las tuplas en la primera fila
    encabezados = [tupla[0] for tupla in array]
    for i, encabezado in enumerate(encabezados):
        hoja.cell(row=1, column=i+2).value = encabezado

    # Escribe el string pasado como parámetro en la primera columna de la segunda fila
    hoja.cell(row=2, column=1).value = string

    # Escribe los segundos elementos de las tuplas en la segunda fila
    valores = [tupla[1] for tupla in array]
    for i, valor in enumerate(valores):
        hoja.cell(row=2, column=i+2).value = valor

    # Guarda el libro de Excel
    libro.save("iter_prompt.xlsx")

#crear_tabla_excel("Let's think step by step", testing('gpt-3.5-turbo', 'tests.dev.v0', "Let's think step by step"))

""" def agregar_fila_excel(tabla_excel, string, array):
    # Carga el libro de Excel existente
    libro = load_workbook(tabla_excel)
    hoja = libro.active

    # Obtiene el número de filas en la hoja
    num_filas = hoja.max_row

    # Obtiene el número de columnas en la hoja
    num_columnas = hoja.max_column

    # Obtiene la siguiente fila disponible
    nueva_fila = num_filas + 1

    # Escribe el string en la primera columna de la nueva fila
    hoja.cell(row=nueva_fila, column=1).value = string

    # Escribe los segundos elementos de las tuplas en las columnas restantes
    for i, tupla in enumerate(array):
        hoja.cell(row=nueva_fila, column=i+2).value = tupla[1]

    # Guarda el libro de Excel con la nueva fila agregada
    libro.save(tabla_excel) """

def agregar_fila_excel(archivo_excel, string, lista_tuplas):
    # Abrir el archivo Excel
    libro_excel = openpyxl.load_workbook(archivo_excel)
    hoja = libro_excel.active

    fila_actual = 3  # Comenzar desde la siguiente fila

    #hoja.cell(row=fila_actual, column=1).value = string
    #hoja.cell(row=fila_actual, column=2).value = lista_tuplas[0][0]
    columna_actual = 7 # Comenzar desde la segunda columna

    for tupla in lista_tuplas:
        i = 0
        while i < len(tupla):
            celda = hoja.cell(row=fila_actual, column=columna_actual)
            celda.value = tupla[i]
            columna_actual += 1
            i += 1
        columna_actual = 7
        fila_actual = fila_actual + 1
        
    # Guardar los cambios en el archivo Excel
    libro_excel.save(archivo_excel)
    libro_excel.close()

#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus11.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus12.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus13.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus14.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus15.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus16.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus17.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus18.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus19.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus20.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus21.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus22.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus23.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus24.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus26.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Let's think step by step."))

#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus11.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus12.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus13.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus14.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus15.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus16.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus17.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus18.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus19.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus20.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus21.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus22.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus23.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/versus24.xlsx', " ", testing('gpt-3.5-turbo', 'versus.dev.v0', "Show step by step the way to arrive to the right answer."))


#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', " ", testing('gpt-3.5-turbo', 'first-letters.dev.v0', ""))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's think step by step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's think step by step"))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Explain your reasoning", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Explain your reasoning"))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's think step by step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's think step by step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's break it down and think through each step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's break it down and think through each step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We should approach this systematically and think step by step to arrive at the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We should approach this systematically and think step by step to arrive at the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Take a moment to carefully consider each step and find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Take a moment to carefully consider each step and find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's analyze the problem and think critically at each stage to find the right solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's analyze the problem and think critically at each stage to find the right solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We'll examine each aspect individually to reach the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We'll examine each aspect individually to reach the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We'll thoroughly scrutinize each option until we find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We'll thoroughly scrutinize each option until we find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's consider each alternative carefully before making a decision", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's consider each alternative carefully before making a decision."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's develop a step-by-step plan to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's develop a step-by-step plan to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's dedicate ourselves to analyzing each element to find the appropriate solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's dedicate ourselves to analyzing each element to find the appropriate solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We need to approach this problem by thinking through each step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We need to approach this problem by thinking through each step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's carefully consider each stage and think step by step in order to arrive at the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's carefully consider each stage and think step by step in order to arrive at the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Taking a systematic approach, let's think through each step to uncover the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Taking a systematic approach, let's think through each step to uncover the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We should break down the problem and think critically at every step to find the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We should break down the problem and think critically at every step to find the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By analyzing each aspect individually, we can determine the right answer step by step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By analyzing each aspect individually, we can determine the right answer step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Adopting a methodical mindset, let's work through each step to reach the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Adopting a methodical mindset, let's work through each step to reach the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We'll thoroughly evaluate each option until we identify the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We'll thoroughly evaluate each option until we identify the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Before making a decision, let's thoughtfully consider each alternative", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Before making a decision, let's thoughtfully consider each alternative."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "To find the right answer, let's develop a clear, step-by-step plan", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "To find the right answer, let's develop a clear, step-by-step plan."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By carefully examining each element, we can uncover the appropriate solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By carefully examining each element, we can uncover the appropriate solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's think as an expert, adpting a methodical minset, step by step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's think as an expert, adopting a methodical minset, step by step to find the right answer"))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's do this step by step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's do this step by step to find the right answer"))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's take a logical approach and think step by step to find the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's take a logical approach and think step by step to find the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We should employ a systematic thought process to uncover the right solution, one step at a time", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We should employ a systematic thought process to uncover the right solution, one step at a time."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By carefully considering each step, we can arrive at the right answer through logical thinking", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By carefully considering each step, we can arrive at the right answer through logical thinking."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's break down the problem into manageable steps and think through each one to reach the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's break down the problem into manageable steps and think through each one to reach the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's carefully and methodically think step by step in order to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's carefully and methodically think step by step in order to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's thoughtfully analyze and think step by step to arrive at the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's thoughtfully analyze and think step by step to arrive at the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's systematically think through each step to uncover the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's systematically think through each step to uncover the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's consciously break down the problem and think step by step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's consciously break down the problem and think step by step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's diligently consider and think step by step to reach the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's diligently consider and think step by step to reach the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's critically and systematically think through each step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's critically and systematically think through each step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's intentionally and progressively think step by step to arrive at the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's intentionally and progressively think step by step to arrive at the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's meticulously analyze and think step by step to uncover the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's meticulously analyze and think step by step to uncover the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's consciously and sequentially think through each step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's consciously and sequentially think through each step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's actively and thoughtfully think step by step in order to reach the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's actively and thoughtfully think step by step in order to reach the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By approaching the problem systematically, we can uncover the right solution step by step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By approaching the problem systematically, we can uncover the right solution step by step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By approaching this problem with a step-by-step mindset, we can find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By approaching this problem with a step-by-step mindset, we can find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "To discover the correct solution, it's important to think sequentially and consider each step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "To discover the correct solution, it's important to think sequentially and consider each step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We should carefully analyze the problem, thinking through each step to reach the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We should carefully analyze the problem, thinking through each step to reach the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Taking a systematic approach, we can uncover the correct answer by considering each step individually", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Taking a systematic approach, we can uncover the correct answer by considering each step individually."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "In order to find the right answer, we must methodically think through each successive step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "In order to find the right answer, we must methodically think through each successive step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By breaking down the problem and thinking step by step, we can arrive at the appropriate solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By breaking down the problem and thinking step by step, we can arrive at the appropriate solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "It's crucial to think sequentially and thoughtfully consider each step to find the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "It's crucial to think sequentially and thoughtfully consider each step to find the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "We can reach the right solution by meticulously thinking through each step of the process", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "We can reach the right solution by meticulously thinking through each step of the process."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "By considering each step in a logical progression, we can uncover the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "By considering each step in a logical progression, we can uncover the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "To arrive at the correct answer, it's necessary to think systematically and evaluate each step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "To arrive at the correct answer, it's necessary to think systematically and evaluate each step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think step by step to find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think step by step to find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Sequentially analyze to reach the correct solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Sequentially analyze to reach the correct solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Consider each step carefully for the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Consider each step carefully for the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Systematically think through to uncover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Systematically think through to uncover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Break it down, think step by step, find the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Break it down, think step by step, find the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Methodically analyze each step for the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Methodically analyze each step for the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Progressively think to arrive at the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Progressively think to arrive at the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Meticulously consider steps for the right solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Meticulously consider steps for the right solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think sequentially, uncover the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think sequentially, uncover the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Evaluate steps to find the appropriate solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Evaluate steps to find the appropriate solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Step by step, find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Step by step, find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Analyze each step for the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Analyze each step for the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think carefully, uncover the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think carefully, uncover the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Systematically consider for the right solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Systematically consider for the right solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Progress through steps, find the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Progress through steps, find the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Sequential thinking reveals the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Sequential thinking reveals the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Thoughtfully evaluate each step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Thoughtfully evaluate each step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Methodically uncover the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Methodically uncover the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Evaluate steps, find the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Evaluate steps, find the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Break it down, find the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Break it down, find the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Let's break down the problem into clear and concise steps to obtain a detailed and precise solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Let's break down the problem into clear and concise steps to obtain a detailed and precise solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Analyze each step to arrive to the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Analyze each step to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think secuencially to find the correct answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think secuencially to find the correct answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Show step by step the way to arrive to the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Show step by step the way to arrive to the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think step by step for the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think step by step for the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Analyze sequentially, find the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Analyze sequentially, find the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Carefully consider each step", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Carefully consider each step."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Systematic thinking leads to the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Systematic thinking leads to the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Progress through steps, uncover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Progress through steps, uncover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Evaluate each step, find the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Evaluate each step, find the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Methodically think to reach the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Methodically think to reach the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Thoughtfully analyze for the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Thoughtfully analyze for the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Break it down, discover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Break it down, discover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Stepwise thinking reveals the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Stepwise thinking reveals the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think step by step, uncover the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think step by step, uncover the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Analyze each stage for the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Analyze each stage for the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Consider steps carefully, find the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Consider steps carefully, find the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Systematically think to reach the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Systematically think to reach the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Progress through steps, discover the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Progress through steps, discover the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Evaluate, step by step, find the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Evaluate, step by step, find the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Methodical thinking leads to the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Methodical thinking leads to the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Thoughtfully assess each step for the right answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Thoughtfully assess each step for the right answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Break it down, find the solution gradually", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Break it down, find the solution gradually."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Stepwise approach reveals the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Stepwise approach reveals the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think step by step, reach the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think step by step, reach the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Analyze each step, find the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Analyze each step, find the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Carefully consider steps for the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Carefully consider steps for the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Systematically think, uncover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Systematically think, uncover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Progress through steps, unveil the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Progress through steps, unveil the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Evaluate each stage, discover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Evaluate each stage, discover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Methodical thinking leads to the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Methodical thinking leads to the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Thoughtfully assess steps for the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Thoughtfully assess steps for the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Break it down, find the solution gradually", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Break it down, find the solution gradually."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Step by step reveals the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Step by step reveals the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Think step by step, unravel the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Think step by step, unravel the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Analyze each stage, discover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Analyze each stage, discover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Carefully consider steps to find the answer", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Carefully consider steps to find the answer."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Systematically think, unveil the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Systematically think, unveil the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Evaluate stages, uncover the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Evaluate stages, uncover the solution."))
#agregar_fila_excel('/Users/camilo.basualdo/Downloads/evals/iter_prompt_v3.xlsx', "Break it down, gradually find the solution", testing('gpt-3.5-turbo', 'first-letters.dev.v0', "Break it down, gradually find the solution."))




""" 
def main():
    parser = get_parser()
    a= sys.argv[1:]
    #args = parser.parse_args(sys.argv[1:])
    a =  ['gpt-3.5-turbo', 'first-letters.dev.v0']
    args = parser.parse_args(a)
    phrase = 'esto es una prueba'
    run(args, phrase)

if __name__ == "__main__":
    main() """
 